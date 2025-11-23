import openpyxl as xl
import json
from investments import *
from expenses import *
from loans import *
from custloans import *
import shutil

summHeaderCol = 2
summAmtCol = summHeaderCol+1 
summLoanNameCol = summAmtCol+1 
summLoanAmtCol = summLoanNameCol+1 

def update_json_by_month(fullDataFile):
    shutil.copy(fullDataFile, 'data_backup.json')
    datafile = open(fullDataFile, "r")
    data = json.load(datafile)

    data['loans'].sort(key=lambda x: int(x['loanEMI']), reverse=True)

    currDet = data['currDet'][0]
    currDate = currDet['date']
    currDate = next_month(currDate)
    data['currDet'][0]['date'] = str(currDate)[0:10]

    loans = data['loans']
    loansLen = len(loans)

    for i in range(0, loansLen):
        loan = data['loans'][i]
        if int(loan['loanEMI']) > 0:
            loan['loanEMI'] = str(int(loan['loanEMI']) - 1)
        else:
            del data['loans'][i]

    datafile.close()

    datafile = open(fullDataFile, "w")
    datafile.write(json.dumps(data, indent=2))
    datafile.truncate()
    datafile.close()

def read_data_file(fullDataFile):

    datafile = open(fullDataFile)
    data = json.load(datafile)

    loans = data['loans']
    loans.sort(key=lambda x: int(x['loanEMI']), reverse=True)

    custloans = data['custloans']

    investments = data['investments']
    investments.sort(key=lambda x: int(x['investDate']))

    expenses = data['expenses']
    expenses.sort(key=lambda x: int(x['expDate']))
    currDet = data['currDet'][0]

    return datafile,currDet,investments,expenses,loans,custloans

def open_sheet(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    #Delete the sheet
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row, col)
            cell.value = ''

    return wb,sheet

def write_header(sheet):
    rowNo=1
    sheet.cell(rowNo, 1).value = 'Date'
    sheet.cell(rowNo, 2).value = 'Expense Type'
    sheet.cell(rowNo, 3).value = 'Amount'
    sheet.cell(rowNo, 4).value = 'EMI Left'
    sheet.cell(rowNo, 5).value = 'Total Amount'
    rowNo+=1

    return rowNo

def write_summary(sheet, rowNo, currDet, bankDet, totInvest, totExp, totLoan, custloans):

    totExpAndLoan = totExp + totLoan + totInvest

    sheet.cell(rowNo, summHeaderCol).value = 'Total Exp'
    sheet.cell(rowNo, summAmtCol).value = totExpAndLoan

    # Write Total EMI Details
    if (totLoan > 0):
        sheet.cell(rowNo, summLoanNameCol).value = 'Total EMI'
        sheet.cell(rowNo, summLoanAmtCol).value = "=sum(" + sheet.cell(2, 5).coordinate + ":" + sheet.cell(rowNo-1, 5).coordinate + ")"
        sheet.cell(rowNo+1, summHeaderCol).value = 'Date'
        sheet.cell(rowNo+1, summAmtCol).value = currDet['date']


    # Write Custom Loan Details
    rowNo+=1
    rowNo,custloansCnt = write_custloan_details(sheet, rowNo, custloans)

    # Write Total Loan Details
    if (totLoan > 0 or custloansCnt > 0):
        sheet.cell(rowNo, summLoanNameCol).value = 'Total Loan'
        sheet.cell(rowNo, summLoanAmtCol).value = "=sum(" + sheet.cell(rowNo-custloansCnt-1, 5).coordinate + ":" + sheet.cell(rowNo-1, 5).coordinate + ")"
        rowNo+=1

        sheet.cell(rowNo, summLoanNameCol).value = 'Year Starting'    
        sheet.cell(rowNo, summLoanAmtCol).value = int(currDet['yearStAmt'])
        rowNo+=1

    rowNo+=1
    sheet.cell(rowNo, summHeaderCol).value = 'Investments'
    sheet.cell(rowNo, summAmtCol).value = totInvest
    totInvestAbsCell = "$" + getCellNo(sheet, rowNo, summAmtCol)[0] + "$" + getCellNo(sheet, rowNo, summAmtCol)[1:]

    rowNo+=1
    sheet.cell(rowNo, summHeaderCol).value = 'Expense'
    sheet.cell(rowNo, summAmtCol).value = totExp
    totExpAbsCell = "$" + getCellNo(sheet, rowNo, summAmtCol)[0] + "$" + getCellNo(sheet, rowNo, summAmtCol)[1:]

    if (totLoan > 0):
        rowNo+=1
        sheet.cell(rowNo, summHeaderCol).value = 'Debt'
        sheet.cell(rowNo, summAmtCol).value = totLoan
        totLoanAbsCell = "$" + getCellNo(sheet, rowNo, summAmtCol)[0] + "$" + getCellNo(sheet, rowNo, summAmtCol)[1:]
        rowNo+=1

    #Bank Section
    #rowNo+=1
    #for bankName in bankDet.keys():
        #sheet.cell(rowNo, summHeaderCol).value = bankName
        #sheet.cell(rowNo, summAmtCol).value = bankDet[bankName]
        #rowNo+=1
    
    rowNo+=1
    return rowNo,totExpAbsCell,totInvestAbsCell

def write_datewise(sheet, rowNo, dateDet):
    sheet.cell(rowNo, 2).value = 'Date-wise'
    sheet.cell(rowNo, 3).value = 'Amount'
    sheet.cell(rowNo, 4).value = 'Commulative'

    sorted_dates=list(dateDet.keys())
    sorted_dates.sort()
    rev_sorted_dates=sorted(sorted_dates, reverse=True)
    commDet = {}
    for date in rev_sorted_dates:
        for in_date in sorted_dates:
            if date <= in_date:
                if commDet.get(date) == None:
                    commDet[date] = dateDet[in_date]
                else:
                    commDet[date] = commDet.get(date) + dateDet[in_date]

    rowNo+=1
    for date in sorted_dates:
        sheet.cell(rowNo, 2).value = date
        sheet.cell(rowNo, 3).value = dateDet[date]
        sheet.cell(rowNo, 4).value = commDet[date]
        rowNo+=1
    
    rowNo+=1
    return rowNo

def write_excel(filename, fullDataFile):

    bankDet = {}
    dateDet = {}

    # Read the data file
    datafile,currDet,investments,expenses,loans,custloans = read_data_file(fullDataFile)

    # Open sheet for writing
    wb,sheet = open_sheet(filename)

    # Write Header Details
    rowNo = write_header(sheet)

    # Write Expense Details
    rowNo, bankDet, dateDet, totExp = write_expense_details(sheet, rowNo, expenses, bankDet, dateDet)

    # Write Investment Details
    rowNo, bankDet, dateDet, totInvest = write_invest_details(sheet, rowNo, investments, bankDet, dateDet)

    # Write Loan Details
    rowNo, bankDet, dateDet, totLoan, maxEMI, totPending = write_loan_details(sheet, rowNo, loans, bankDet, dateDet)

    # Write Summary
    rowNo,totExpAbsCell,totInvestAbsCell = write_summary(sheet, rowNo, currDet, bankDet, totInvest, totExp, totLoan, custloans)

    # Write Date-wise Summary
    rowNo = write_datewise(sheet, rowNo, dateDet)

    # Write Loan EMI Details
    totalLoanAndOthers = totExp + totInvest
    if (maxEMI > 0):
        write_loan_emi_details(sheet, currDet, loans, maxEMI, totExpAbsCell, totInvestAbsCell, totPending, totalLoanAndOthers, rowNo)

    wb.save(filename)

    datafile.close()