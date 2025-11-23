import openpyxl as xl
import json
import os
from datetime import datetime, timedelta
import calendar

def write_invest_details(sheet, rowNo, investments, bankDet, dateDet):

    totInvest = 0
    for invest in investments:
        investAmt = float(invest['investAmt'])
        investDate = int(invest['investDate'])
        sheet.cell(rowNo, 1).value = investDate
        sheet.cell(rowNo, 2).value = invest['investName']
        sheet.cell(rowNo, 3).value = investAmt
        sheet.cell(rowNo, 4).value = 0
        sheet.cell(rowNo, 5).value = 0
        rowNo = rowNo + 1

        totInvest = totInvest + investAmt
        if bankDet.get(invest['investBank']) == None:
            bankDet[invest['investBank']] = investAmt
        else:
            bankDet[invest['investBank']] = bankDet.get(invest['investBank']) + investAmt

        if dateDet.get(investDate) == None:
            dateDet[investDate] = investAmt
        else:
            dateDet[investDate] = dateDet.get(investDate) + investAmt

    return rowNo, bankDet, dateDet, totInvest