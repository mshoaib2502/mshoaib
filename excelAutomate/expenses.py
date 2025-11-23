import openpyxl as xl
import json
import os
from datetime import datetime, timedelta
import calendar

def write_expense_details(sheet, rowNo, expenses, bankDet, dateDet):

    totExp = 0

    for exp in expenses:
        expAmt = float(exp['expAmt'])
        expDate = int(exp['expDate'])
        sheet.cell(rowNo, 1).value = expDate
        sheet.cell(rowNo, 2).value = exp['expName']
        sheet.cell(rowNo, 3).value = expAmt
        sheet.cell(rowNo, 4).value = 0
        sheet.cell(rowNo, 5).value = 0
        rowNo = rowNo + 1

        totExp = totExp + expAmt
        if bankDet.get(exp['expBank']) == None:
            bankDet[exp['expBank']] = expAmt
        else:
            bankDet[exp['expBank']] = bankDet.get(exp['expBank']) + expAmt

        if dateDet.get(expDate) == None:
            dateDet[expDate] = expAmt
        else:
            dateDet[expDate] = dateDet.get(expDate) + expAmt

    return rowNo, bankDet, dateDet, totExp